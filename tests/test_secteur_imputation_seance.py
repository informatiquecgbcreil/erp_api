"""Secteur d'imputation choisi séance par séance.

Une séance peut être animée pour un autre secteur que celui de son atelier
(« j'émarge pour les Familles ») : elle sort alors de mes statistiques et
entre dans celles du secteur choisi, sans que personne ne perde l'accès.
"""
import datetime as dt
import uuid
from io import BytesIO

from openpyxl import load_workbook


def _client_secteur(app, *, email, secteur, role="responsable_secteur"):
    """Compte cloisonné sur un secteur (pas de portée globale)."""
    from app.extensions import db
    from app.models import Role, User
    with app.app_context():
        u = User.query.filter_by(email=email).first()
        if u is None:
            u = User(email=email, nom=email.split("@")[0], secteur_assigne=secteur)
            u.set_password("pw-test-123")
            u.roles.append(Role.query.filter_by(code=role).first())
            db.session.add(u)
            db.session.commit()
    c = app.test_client()
    r = c.post("/", data={"email": email, "password": "pw-test-123"})
    assert r.status_code == 302, "connexion échouée"
    return c


def _declarer_secteurs(app, *labels):
    """Le référentiel des secteurs fait foi pour valider une imputation :
    un libellé absent est refusé (voir _secteur_imputation_depuis_formulaire)."""
    from app.extensions import db
    from app.models import Secteur
    with app.app_context():
        for label in labels:
            if Secteur.query.filter_by(label=label).first() is None:
                db.session.add(Secteur(code=label.lower(), label=label, is_active=True))
        db.session.commit()


def _atelier(app, *, secteur, nom):
    from app.extensions import db
    from app.models import AtelierActivite
    _declarer_secteurs(app, secteur)
    with app.app_context():
        at = AtelierActivite(nom=nom, secteur=secteur, type_atelier="COLLECTIF",
                             capacite_defaut=10, is_active=True)
        db.session.add(at)
        db.session.commit()
        return at.id


def _seance(app, atelier_id, *, secteur_impute, jour=10, presences=0):
    from app.extensions import db
    from app.models import Participant, PresenceActivite, SessionActivite
    annee = dt.date.today().year
    with app.app_context():
        s = SessionActivite(
            atelier_id=atelier_id, secteur=secteur_impute, session_type="COLLECTIF",
            date_session=dt.date(annee, 3, jour), heure_debut="14:00", heure_fin="16:00",
            capacite=10, statut="realisee",
        )
        db.session.add(s)
        db.session.flush()
        for i in range(presences):
            p = Participant(nom=f"P{uuid.uuid4().hex[:6]}", prenom=str(i), genre="Femme")
            db.session.add(p)
            db.session.flush()
            db.session.add(PresenceActivite(session_id=s.id, participant_id=p.id))
        db.session.commit()
        return s.id


# ---------------------------------------------------------------------------
# Saisie
# ---------------------------------------------------------------------------

def test_creation_seance_avec_secteur_choisi(app, admin_client):
    from app.extensions import db
    from app.models import SessionActivite

    tag = uuid.uuid4().hex[:6]
    secteur_atelier, secteur_cible = f"Num{tag}", f"Fam{tag}"
    _declarer_secteurs(app, secteur_cible)
    atelier_id = _atelier(app, secteur=secteur_atelier, nom=f"Atelier{tag}")

    r = admin_client.post(
        f"/activite/atelier/{atelier_id}/session/new",
        data={
            "date_session": dt.date(dt.date.today().year, 3, 12).isoformat(),
            "heure_debut": "14:00", "heure_fin": "16:00", "capacite": "10",
            "secteur_impute": secteur_cible,
        },
    )
    assert r.status_code == 302
    with app.app_context():
        s = SessionActivite.query.filter_by(atelier_id=atelier_id).one()
        assert s.secteur == secteur_cible, "le secteur choisi doit être retenu"


def test_secteur_inconnu_retombe_sur_celui_de_l_atelier(app, admin_client):
    """Un libellé fantaisiste ne doit jamais envoyer les stats dans le vide."""
    from app.extensions import db
    from app.models import SessionActivite

    tag = uuid.uuid4().hex[:6]
    secteur_atelier = f"Num{tag}"
    atelier_id = _atelier(app, secteur=secteur_atelier, nom=f"Atelier{tag}")

    r = admin_client.post(
        f"/activite/atelier/{atelier_id}/session/new",
        data={
            "date_session": dt.date(dt.date.today().year, 3, 12).isoformat(),
            "heure_debut": "14:00", "heure_fin": "16:00",
            "secteur_impute": "Secteur qui n'existe pas",
        },
    )
    assert r.status_code == 302
    with app.app_context():
        s = SessionActivite.query.filter_by(atelier_id=atelier_id).one()
        assert s.secteur == secteur_atelier


def test_creation_en_serie_applique_le_secteur(app, admin_client):
    from app.extensions import db
    from app.models import SessionActivite

    tag = uuid.uuid4().hex[:6]
    secteur_atelier, secteur_cible = f"Num{tag}", f"Fam{tag}"
    _declarer_secteurs(app, secteur_cible)
    atelier_id = _atelier(app, secteur=secteur_atelier, nom=f"Atelier{tag}")
    annee = dt.date.today().year

    r = admin_client.post(
        f"/activite/atelier/{atelier_id}/sessions/bulk",
        data={
            "date_debut": dt.date(annee, 3, 2).isoformat(),
            "date_fin": dt.date(annee, 3, 16).isoformat(),
            "weekday": "0", "heure_debut": "14:00", "heure_fin": "16:00",
            "secteur_impute": secteur_cible,
        },
    )
    assert r.status_code == 302
    with app.app_context():
        seances = SessionActivite.query.filter_by(atelier_id=atelier_id).all()
        assert seances, "des séances doivent être créées"
        assert {s.secteur for s in seances} == {secteur_cible}


def test_changement_de_secteur_trace_dans_le_journal(app, admin_client):
    from app.extensions import db
    from app.models import SessionActivite, SessionScheduleEditLog

    tag = uuid.uuid4().hex[:6]
    secteur_atelier, secteur_cible = f"Num{tag}", f"Fam{tag}"
    _declarer_secteurs(app, secteur_cible)
    atelier_id = _atelier(app, secteur=secteur_atelier, nom=f"Atelier{tag}")
    session_id = _seance(app, atelier_id, secteur_impute=secteur_atelier)

    with app.app_context():
        jour = db.session.get(SessionActivite, session_id).date_session

    r = admin_client.post(
        f"/activite/session/{session_id}/edit-schedule",
        data={
            "date_session": jour.isoformat(), "heure_debut": "14:00", "heure_fin": "16:00",
            "capacite": "10", "secteur_impute": secteur_cible,
            "edit_reason": "Animée par le collègue des familles",
        },
    )
    assert r.status_code == 302
    with app.app_context():
        s = db.session.get(SessionActivite, session_id)
        assert s.secteur == secteur_cible
        log = SessionScheduleEditLog.query.filter_by(session_id=session_id).one()
        assert secteur_atelier in log.reason and secteur_cible in log.reason


# ---------------------------------------------------------------------------
# Statistiques : c'est le secteur de la séance qui décide
# ---------------------------------------------------------------------------

def test_seance_imputee_ailleurs_sort_de_mes_stats(app):
    """Le cœur de la demande : ma séance animée pour un autre secteur
    quitte mes statistiques et entre dans les siennes."""
    from app.statsimpact.engine import compute_volume_activity_stats, normalize_filters

    tag = uuid.uuid4().hex[:6]
    mon_secteur, autre_secteur = f"Num{tag}", f"Fam{tag}"
    atelier_id = _atelier(app, secteur=mon_secteur, nom=f"Atelier{tag}")
    _seance(app, atelier_id, secteur_impute=mon_secteur, jour=10, presences=2)
    _seance(app, atelier_id, secteur_impute=autre_secteur, jour=17, presences=3)

    annee = dt.date.today().year
    args = {"date_from": f"{annee}-01-01", "date_to": f"{annee}-12-31"}

    _client_secteur(app, email=f"moi-{tag}@ex.org", secteur=mon_secteur)
    _client_secteur(app, email=f"autre-{tag}@ex.org", secteur=autre_secteur)

    # On calcule dans le contexte de chaque utilisateur (le moteur lit current_user).
    with app.test_request_context(query_string=args):
        from flask_login import login_user
        from app.models import User
        from app.extensions import db

        u_moi = User.query.filter_by(email=f"moi-{tag}@ex.org").first()
        login_user(u_moi)
        stats_moi = compute_volume_activity_stats(normalize_filters(dict(args), user=u_moi))

    with app.test_request_context(query_string=args):
        from flask_login import login_user
        from app.models import User

        u_autre = User.query.filter_by(email=f"autre-{tag}@ex.org").first()
        login_user(u_autre)
        stats_autre = compute_volume_activity_stats(normalize_filters(dict(args), user=u_autre))

    assert stats_moi["kpi"]["sessions"] == 1, "seule ma séance reste chez moi"
    assert stats_moi["kpi"]["presences"] == 2
    assert stats_autre["kpi"]["sessions"] == 1, "la séance animée pour eux compte chez eux"
    assert stats_autre["kpi"]["presences"] == 3


def test_export_detaille_suit_l_imputation(app, admin_client):
    """L'atelier apparaît dans l'export du secteur bénéficiaire, avec les
    seules séances qui lui sont imputées."""
    tag = uuid.uuid4().hex[:6]
    mon_secteur, autre_secteur = f"Num{tag}", f"Fam{tag}"
    atelier_id = _atelier(app, secteur=mon_secteur, nom=f"Atelier{tag}")
    _seance(app, atelier_id, secteur_impute=mon_secteur, jour=10, presences=2)
    _seance(app, atelier_id, secteur_impute=autre_secteur, jour=17, presences=3)

    def _ligne(secteur):
        r = admin_client.get(f"/stats-impact/magatomatique.xlsx?export_mode=complete&secteur={secteur}")
        assert r.status_code == 200
        ws = load_workbook(BytesIO(r.data))["Synthese"]
        entetes = [c.value for c in ws[3]]
        for row in ws.iter_rows(min_row=4, values_only=True):
            if row[1] == f"Atelier{tag}":
                return dict(zip(entetes, row))
        return None

    chez_moi = _ligne(mon_secteur)
    assert chez_moi is not None
    assert chez_moi["Nb séances réelles"] == 1
    assert chez_moi["Nb présences totales"] == 2

    chez_eux = _ligne(autre_secteur)
    assert chez_eux is not None, "l'atelier doit apparaître dans l'export du secteur bénéficiaire"
    assert chez_eux["Nb séances réelles"] == 1
    assert chez_eux["Nb présences totales"] == 3


# ---------------------------------------------------------------------------
# Accès : personne ne perd la main
# ---------------------------------------------------------------------------

def test_acces_conserve_des_deux_cotes(app):
    """Le secteur porteur de l'atelier garde l'accès, le secteur bénéficiaire
    l'obtient : une imputation croisée ne rend jamais une séance orpheline."""
    tag = uuid.uuid4().hex[:6]
    secteur_atelier, secteur_cible = f"Num{tag}", f"Fam{tag}"
    atelier_id = _atelier(app, secteur=secteur_atelier, nom=f"Atelier{tag}")
    session_id = _seance(app, atelier_id, secteur_impute=secteur_cible)

    porteur = _client_secteur(app, email=f"porteur-{tag}@ex.org", secteur=secteur_atelier)
    r = porteur.get(f"/activite/session/{session_id}/emargement")
    assert r.status_code == 200, "le secteur de l'atelier garde l'accès"

    beneficiaire = _client_secteur(app, email=f"benef-{tag}@ex.org", secteur=secteur_cible)
    r = beneficiaire.get(f"/activite/session/{session_id}/emargement")
    assert r.status_code == 200, "le secteur d'imputation y a accès"

    etranger = _client_secteur(app, email=f"tiers-{tag}@ex.org", secteur=f"Autre{tag}")
    r = etranger.get(f"/activite/session/{session_id}/emargement", follow_redirects=False)
    assert r.status_code == 302, "un secteur non concerné reste dehors"
