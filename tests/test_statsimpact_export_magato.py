"""Export Magatomatique détaillé (1 feuille par atelier).

Deux exigences de fiabilité côté bilans financeurs :
- une séance mise à la corbeille ne doit plus compter nulle part ;
- le genre doit être lisible par personne ET agrégé par atelier.
"""
import datetime as dt
import uuid
from io import BytesIO

from openpyxl import load_workbook


def _jeu_de_donnees(app, *, tag, avec_corbeille=False):
    """Un atelier, deux séances, trois participants émargés.

    Renvoie (secteur, nom_atelier). Quand ``avec_corbeille`` est vrai, une
    troisième séance (avec sa présence) est mise à la corbeille : elle ne
    doit apparaître dans aucun total.
    """
    from app.extensions import db
    from app.models import AtelierActivite, Participant, PresenceActivite, SessionActivite

    secteur = f"Sect{tag}"
    with app.app_context():
        atelier = AtelierActivite(
            nom=f"Atelier{tag}", secteur=secteur, type_atelier="COLLECTIF",
            capacite_defaut=10, is_active=True,
        )
        db.session.add(atelier)
        db.session.flush()

        # Dates dans l'année civile courante (période par défaut des exports).
        annee = dt.date.today().year
        seances = []
        for jour in (10, 17):
            s = SessionActivite(
                atelier_id=atelier.id, secteur=secteur, session_type="COLLECTIF",
                date_session=dt.date(annee, 3, jour), heure_debut="14:00", heure_fin="16:00",
                capacite=10, statut="realisee",
            )
            db.session.add(s)
            seances.append(s)

        participants = [
            Participant(nom=f"Femme{tag}", prenom="Alice", genre="Femme",
                        date_naissance=dt.date(1990, 5, 4)),
            Participant(nom=f"Homme{tag}", prenom="Bruno", genre="Homme",
                        date_naissance=dt.date(1985, 2, 20)),
            Participant(nom=f"Nr{tag}", prenom="Camille", genre=None,
                        date_naissance=dt.date(2000, 9, 9)),
        ]
        db.session.add_all(participants)
        db.session.flush()

        for s in seances:
            for p in participants:
                db.session.add(PresenceActivite(session_id=s.id, participant_id=p.id))

        if avec_corbeille:
            fantome = SessionActivite(
                atelier_id=atelier.id, secteur=secteur, session_type="COLLECTIF",
                date_session=dt.date(annee, 3, 24), heure_debut="14:00", heure_fin="16:00",
                capacite=10, statut="realisee",
                is_deleted=True, deleted_at=dt.datetime(annee, 3, 25),
            )
            db.session.add(fantome)
            db.session.flush()
            for p in participants:
                db.session.add(PresenceActivite(session_id=fantome.id, participant_id=p.id))

        db.session.commit()
        return secteur, atelier.nom


def _exporter(admin_client, secteur):
    r = admin_client.get(f"/stats-impact/magatomatique.xlsx?export_mode=complete&secteur={secteur}")
    assert r.status_code == 200
    return load_workbook(BytesIO(r.data))


def _ligne_synthese(wb, nom_atelier):
    ws = wb["Synthese"]
    entetes = [c.value for c in ws[3]]
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[1] == nom_atelier:
            return dict(zip(entetes, row))
    raise AssertionError(f"Atelier {nom_atelier} absent de la synthèse")


def test_export_ignore_les_seances_en_corbeille(app, admin_client):
    """Régression : une séance supprimée (corbeille) gonflait séances,
    heures et présences de l'export détaillé."""
    tag = uuid.uuid4().hex[:6]
    secteur, nom_atelier = _jeu_de_donnees(app, tag=tag, avec_corbeille=True)

    ligne = _ligne_synthese(_exporter(admin_client, secteur), nom_atelier)

    # 2 séances valides (la 3e est à la corbeille), 3 présences chacune.
    assert ligne["Nb séances prévisionnelles"] == 2
    assert ligne["Nb séances réelles"] == 2
    assert ligne["Nb présences totales"] == 6
    assert ligne["Nb inscrits (uniques)"] == 3
    assert ligne["Nb heures réalisées"] == 4.0


def test_export_compte_les_seances_valides(app, admin_client):
    """Contrôle inverse : sans corbeille, les mêmes chiffres sortent —
    le filtre ne mange pas de séances légitimes."""
    tag = uuid.uuid4().hex[:6]
    secteur, nom_atelier = _jeu_de_donnees(app, tag=tag)

    ligne = _ligne_synthese(_exporter(admin_client, secteur), nom_atelier)
    assert ligne["Nb séances réelles"] == 2
    assert ligne["Nb présences totales"] == 6


def test_synthese_porte_la_repartition_par_genre(app, admin_client):
    tag = uuid.uuid4().hex[:6]
    secteur, nom_atelier = _jeu_de_donnees(app, tag=tag)

    ligne = _ligne_synthese(_exporter(admin_client, secteur), nom_atelier)
    assert ligne["Femmes"] == 1
    assert ligne["Hommes"] == 1
    assert ligne["Autre / NR"] == 1


def test_feuille_atelier_genre_par_personne_et_agrege(app, admin_client):
    tag = uuid.uuid4().hex[:6]
    secteur, nom_atelier = _jeu_de_donnees(app, tag=tag)
    wb = _exporter(admin_client, secteur)

    ws = wb[nom_atelier]
    lignes = [[c.value for c in row] for row in ws.iter_rows()]

    # Bloc de répartition agrégé
    bloc = {r[0]: r[1] for r in lignes if r and r[0] in {"Femmes", "Hommes", "Autre / non renseigné"}}
    assert bloc == {"Femmes": 1, "Hommes": 1, "Autre / non renseigné": 1}

    # Colonne Genre dans la matrice, avec un libellé normalisé par personne
    entetes_matrice = next(r for r in lignes if r and r[0] == "Nom")
    assert entetes_matrice[:6] == ["Nom", "Prénom", "Genre", "Âge", "Ville", "Quartier"]
    idx_genre = 2
    genres = {r[0]: r[idx_genre] for r in lignes if r and r[0] and str(r[0]).endswith(tag)}
    assert genres[f"Femme{tag}"] == "Femme"
    assert genres[f"Homme{tag}"] == "Homme"
    assert genres[f"Nr{tag}"] == "Non renseigné"

    # Les croix d'émargement restent alignées malgré la colonne ajoutée :
    # 2 séances émargées pour chaque personne.
    for r in lignes:
        if r and r[0] and str(r[0]).endswith(tag) and r[idx_genre]:
            assert [c for c in r[6:] if c == "1"] == ["1", "1"]


def test_export_fidele_reste_disponible(app, admin_client):
    """L'autre export (statsimpact fidèle) n'est pas cassé par les ajouts."""
    tag = uuid.uuid4().hex[:6]
    secteur, _ = _jeu_de_donnees(app, tag=tag)
    r = admin_client.get(f"/stats-impact/export-fidele.xlsx?secteur={secteur}")
    assert r.status_code == 200
    wb = load_workbook(BytesIO(r.data))
    assert "Publics" in wb.sheetnames
